def update_attendance_excel():
    import openpyxl
    from openpyxl import Workbook
    from datetime import datetime
    import os
    from firebase_admin import db

    # Correct Excel file path and format
    excel_file = "attendance_history.xlsx"

    def get_attendance_from_firebase():
        ref = db.reference('/attendee')
        data = ref.get()
        return data or {}

    def update_excel(data):
        today = datetime.now()
        today_str = today.strftime("%d-%m-%Y")  # Full date for clarity
        sheet_name = today.strftime("%B_%Y")  # e.g., July_2025

        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
            sheet.append(["ID", "Name"])  # Header row

        # Build header index
        headers = [cell.value for cell in sheet[1]]
        if today_str in headers:
            col_index = headers.index(today_str) + 1
        else:
            col_index = len(headers) + 1
            sheet.cell(row=1, column=col_index, value=today_str)

        # Build existing row map
        row_map = {sheet.cell(row=i, column=1).value: i for i in range(2, sheet.max_row + 1)}

        for uid, info in data.items():
            name = info.get("name", "")
            last_time = info.get("last_attendance_time", "")
            attendance_status = "A"

            try:
                last_date = datetime.strptime(last_time, "%Y-%m-%d %H:%M:%S").date()
                if last_date == today.date():
                    attendance_status = "P"
            except Exception as e:
                print(f"Error parsing date for {uid}: {e}")

            if uid in row_map:
                row = row_map[uid]
            else:
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=uid)
                sheet.cell(row=row, column=2, value=name)

            sheet.cell(row=row, column=col_index, value=attendance_status)

        wb.save(excel_file)
        wb.close()
        print(f"Attendance updated in {excel_file} for {sheet_name}.")

    data = get_attendance_from_firebase()
    update_excel(data)
